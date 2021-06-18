from os import getcwd
from os.path import (isfile,
                     join)
from copy import deepcopy
from datetime import date
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment
from win32com.client import Dispatch
from pyodbc import connect as pycon
from sqlite3 import connect as sqcon
from _constants import (ACCESS_GROUPS_COLS,
                        ACCESS_STAND_COLS,
                        ACCESS_TREE_COLS,
                        SQL_GROUPS_COLS,
                        SQL_STAND_COLS,
                        SQL_TREE_COLS,
                        GROUPS_DEFAULTS,
                        extension_check)


class FVS(object):
    def __init__(self):
        """The FVS class will create FVS-formatted databases for use in FVS. FVS is a software made by the US Forest Service to
           run models, inventories and simulations on different forest stands. The three types of databases the FVS class creates
           are Microsoft Access, Microsoft Excel and SQLite

           If you would like a blank database to manually input your data, you can simply instantiate the FVS class and call either

           fvs.access_db(filename, directory=yourdirectory (optional), blank_db=True)
           fvs.excel_db(filename, directory=yourdirectory (optional), blank_db=True) OR
           fvs.sqlite_db(filename, directory=yourdirectory (optional), blank_db=True)

           If the directory argument is omitted, the default directory is the current working directory.
           The 'blank_db' argument has to be set to True for blank databases (its default is False)

           If you would like to create the databases from a Stand Class, you can instantiate the FVS class and call the
           fvs.set_stand method. There are seven required arguments, these are data that the FVS software needs for its models, they include

           stand (the Stand Class)
           variant (str) [These are two-letter acronyms for different eco-regions throughout the US]
           forest_code (int) [This is the forest code for the closest national forest to your stand]
           region (int) [This is the region code for the forest service region your stand is in]
           stand_age (int) [The approximate age of the stand]
           site_species (str) [The species acronym for the species being used to measure site index]
           site_index (int) [The site index of the stand]

           kwargs can be used to add additional data to your stand. To see the list of the other stand data call the
           fvs.show_stand_args method, use this console printout to find the index of your desired argument(s),
           then use the list index to set your argument to fill in the kwargs,
           for example fvs.set_stand(... fvs.stand_args[4]=22)

           After setting the stand in the FVS class, you can then create or append your databases by calling the database-specific methods
           listed above"""

        self.stand_args = [i[0] for i in ACCESS_STAND_COLS[1]]
        self.stand = None
        self.stand_fvs = None
        self.tree_fvs = None

    def show_stand_args(self):
        """Shows the optional stand data and their indices within self.stand_args"""
        format_space = 20
        for i, arg in enumerate(self.stand_args):
            print(f'{arg + (" " * (format_space - len(arg)))}\t Index: {i}')

    def set_stand(self, stand, variant: str, forest_code: int, region: int, stand_age: int, site_species: str, site_index: int, **kwargs):
        """Extracts the stand and tree data from the Stand class and updates the stand_fvs and tree_fvs dictionaries which will be used
           to create or update the databases"""
        self.stand = deepcopy(stand)
        self.stand_fvs = {i[0]: None for i in ACCESS_STAND_COLS[1]}
        self.tree_fvs = {}

        if self.stand.inv_date:
            inv_year = self.stand.inv_date.year
        else:
            inv_year = date.today().year

        self.stand_fvs['Stand_ID'] = self.stand.name
        self.stand_fvs['Variant'] = variant
        self.stand_fvs['Inv_Year'] = inv_year
        self.stand_fvs['Groups'] = 'All_Stands'
        self.stand_fvs['Region'] = region
        self.stand_fvs['Forest'] = forest_code
        self.stand_fvs['Age'] = stand_age
        self.stand_fvs['Basal_Area_Factor'] = self.stand.plot_factor
        self.stand_fvs['Brk_DBH'] = 0
        self.stand_fvs['Num_Plots'] = len(self.stand.plots)
        self.stand_fvs['Site_Species'] = site_species
        self.stand_fvs['Site_Index'] = site_index

        row = 1
        for i, plot in enumerate(self.stand.plots):
            pnum = i + 1
            std_plt = f'{self.stand.name}_{pnum}'
            for j, tree in enumerate(plot.trees):
                self.tree_fvs[row] = {i[0]: None for i in ACCESS_TREE_COLS[1]}
                self.tree_fvs[row]['Stand_ID'] = self.stand.name
                self.tree_fvs[row]['StandPlot_ID'] = std_plt
                self.tree_fvs[row]['Plot_ID'] = pnum
                self.tree_fvs[row]['Tree_ID'] = j + 1
                self.tree_fvs[row]['History'] = 1
                self.tree_fvs[row]['Species'] = tree.species
                self.tree_fvs[row]['DBH'] = tree.dbh
                self.tree_fvs[row]['Ht'] = tree.height
                row += 1

        for key in kwargs:
            if key in self.stand_fvs:
                self.stand_fvs[key] = kwargs[key]

    def access_db(self, filename: str, directory: str = None, blank_db: bool = False):
        """Creates or updates an FVS-formatted Microsoft Access Database (.accdb)"""
        db_save = extension_check(filename, '.accdb')

        if directory:
            dir_ = directory
            db_path = join(dir_, db_save)
        else:
            dir_ = getcwd()
            db_path = join(dir_, db_save)

        db_exists = True

        if not isfile(db_path):
            self._create_access_db(db_path)
            db_exists = False

        drive = '{Microsoft Access Driver (*.mdb, *.accdb)}'
        connection_text = r'DRIVER={driver};DBQ={filename};'.format(driver=drive, filename=db_path)
        con = pycon(connection_text)
        cur = con.cursor()

        if not db_exists:
            sql = f"""INSERT INTO FVS_GroupAddFilesAndKeywords (Groups, FVSKeywords) VALUES (?, ?)"""
            cur.execute(sql, [GROUPS_DEFAULTS[0], GROUPS_DEFAULTS[1].format(DB_NAME=db_save)])
            con.commit()

        if not blank_db:
            self._insert_sql(con, cur)
        con.close()

        self._create_loc_file(db_save, dir_)

    def sqlite_db(self, filename: str, directory: str = None, blank_db: bool = False):
        """Creates or updates an FVS-formatted SQLite Database (.db)"""
        db_save = extension_check(filename, '.db')

        if directory:
            db_path = join(directory, db_save)
        else:
            db_path = join(getcwd(), db_save)

        if not isfile(db_path):
            con, cur = self._create_sqlite_db(db_path, db_save)
        else:
            con = sqcon(db_path)
            cur = con.cursor()

        if not blank_db:
            self._insert_sql(con, cur)

        con.close()

    def excel_db(self, filename: str, directory: str = None, blank_db: bool = False):
        """Creates or updates an FVS-formatted Microsoft Excel Database (.xlsx)"""
        db_save = extension_check(filename, '.xlsx')

        if directory:
            db_path = join(directory, db_save)
        else:
            db_path = join(getcwd(), db_save)

        if not isfile(db_path):
            wb = self._create_excel_db(db_save)
        else:
            wb = load_workbook(db_path)

        if not blank_db:
            self._insert_excel(wb)

        wb.save(db_path)

    def _create_access_db(self, db_path:  str):
        """If the file does not exist, this method is called to construct the Access database, used internally"""
        access = Dispatch("Access.Application")
        access_engine = access.DBEngine
        access_workspace = access_engine.Workspaces(0)
        access_language = ';LANGID=0x0409;CP=1252;COUNTRY=0'
        access_db = access_workspace.CreateDatabase(db_path, access_language, 64)

        for table in [ACCESS_GROUPS_COLS, ACCESS_STAND_COLS, ACCESS_TREE_COLS]:
            sql = f"""CREATE TABLE {table[0]} ({', '.join([f'{i[0]} {i[1]}' for i in table[1]])});"""
            access_db.Execute(sql)

    def _create_sqlite_db(self, db_path: str, db_save: str):
        """If the file does not exist, this method is called to construct the SQLite database, used internally"""
        con = sqcon(db_path)
        cur = con.cursor()
        for i, table in enumerate([SQL_GROUPS_COLS, SQL_STAND_COLS, SQL_TREE_COLS]):
            sql = f'CREATE TABLE {table[0]} (' + ', '.join([f'{i[0]} {i[1]}' for i in table[1]]) + ');'
            cur.execute(sql)

        sql = f"""INSERT INTO FVS_GroupAddFilesAndKeywords (Groups, FVSKeywords) VALUES (?, ?)"""
        cur.execute(sql, [GROUPS_DEFAULTS[0], GROUPS_DEFAULTS[1].format(DB_NAME=db_save)])
        con.commit()
        return con, cur

    def _create_excel_db(self, db_save: str):
        """If the file does not exist, this method is called to construct the Excel database, used internally"""
        wb = Workbook()
        for i, table in enumerate([ACCESS_GROUPS_COLS, ACCESS_STAND_COLS, ACCESS_TREE_COLS]):
            ws = wb.create_sheet(table[0])
            for j, col in enumerate(table[1]):
                ws.cell(1, j + 1).value = col[0]
            if i == 0:
                ws.cell(2, 1).value = GROUPS_DEFAULTS[0]
                ws.cell(2, 3).value = GROUPS_DEFAULTS[1].format(DB_NAME=db_save)
                ws.cell(2, 3).alignment = Alignment(wrapText=True)

        for sheet in wb.sheetnames:
            if sheet not in [ACCESS_GROUPS_COLS[0], ACCESS_STAND_COLS[0], ACCESS_TREE_COLS[0]]:
                wb.remove(wb[sheet])
        return wb

    def _insert_sql(self, connection, cursor):
        """Inserts the stand and tree FVS data into the Access or SQLite database"""
        stand_cols = [col for col in self.stand_fvs if self.stand_fvs[col]]
        stand_vals = [self.stand_fvs[col] for col in self.stand_fvs if self.stand_fvs[col]]
        stand_sql = f"""INSERT INTO FVS_StandInit ({', '.join(stand_cols)}) VALUES ({', '.join(['?' for _ in stand_cols])})"""
        cursor.execute(stand_sql, stand_vals)

        for row in self.tree_fvs:
            tree_cols = [col for col in self.tree_fvs[row] if self.tree_fvs[row][col]]
            tree_vals = [self.tree_fvs[row][col] for col in self.tree_fvs[row] if self.tree_fvs[row][col]]
            tree_sql = f"""INSERT INTO FVS_TreeInit ({', '.join(tree_cols)}) VALUES ({', '.join(['?' for _ in tree_cols])})"""

            cursor.execute(tree_sql, tree_vals)
        connection.commit()

    def _insert_excel(self, workbook):
        """Inserts the stand and tree FVS data into the Excel database"""
        stand_keys = list(self.stand_fvs)
        stand_cols = [col for col in self.stand_fvs if self.stand_fvs[col]]
        stand_idxs = [stand_keys.index(col) + 1 for col in stand_cols]
        stand_vals = [self.stand_fvs[col] for col in self.stand_fvs if self.stand_fvs[col]]

        ws = workbook['FVS_StandInit']
        next_row = ws.max_row + 1
        for idx, val in zip(stand_idxs, stand_vals):
            ws.cell(next_row, idx).value = val

        ws = workbook['FVS_TreeInit']
        next_row = ws.max_row + 1
        tree_keys = list(self.tree_fvs[1])
        for row in self.tree_fvs:
            tree_cols = [col for col in self.tree_fvs[row] if self.tree_fvs[row][col]]
            tree_idxs = [tree_keys.index(col) + 1 for col in tree_cols]
            tree_vals = [self.tree_fvs[row][col] for col in self.tree_fvs[row] if self.tree_fvs[row][col]]
            for idx, val in zip(tree_idxs, tree_vals):
                ws.cell(next_row, idx).value = val
            next_row += 1

    def _create_loc_file(self, db_save, directory):
        """In the FVS software (legacy versions), a Suppose.loc file needs to be made or modified with the Access database,
           this file is created in the same directory as the database"""
        loc_path = join(directory, 'Suppose.loc')
        fill_text = f'{loc_path[0]} "{db_save[0:-6]}" {db_save}'

        if isfile(loc_path):
            in_loc = True
            with open(loc_path, 'r') as f:
                lines = [i.strip('\n') for i in f.readlines()]
                if fill_text not in lines:
                    in_loc = False

            if not in_loc:
                with open(loc_path, 'a') as f:
                    f.write('\n' + fill_text)
        else:
            with open(loc_path, 'w') as f:
                f.write(fill_text)







if __name__ == '__main__':
    from stand import Stand

    stand = Stand('OK2', 46.94)
    stand.from_csv_full('../example_csv_and_xlsx/Example_CSV_full.csv')
    fvs = FVS()
    fvs.show_stand_args()
    fvs.set_stand(stand, 'PN', 612, 6, 50, 'DF', 120)
    fvs.access_db('spam')

