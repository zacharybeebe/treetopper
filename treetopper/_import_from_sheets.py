from csv import reader
from openpyxl import load_workbook
from statistics import mean
from treetopper._constants import (
    SHEET_ROW_COL_CONV,
    SHEET_FULL_LOG_CONV
)
from treetopper._exceptions import ImportSheetError
from treetopper._utils import (
    get_extension,
    is_err_num
)


def read_csv(file, stand_name):
    rows = []
    with open(file, 'r') as csv_file:
        csv_read = reader(csv_file)
        next(csv_read)
        for row in csv_read:
            if row[0].upper() == stand_name.upper():
                rows.append(row)
    if not rows:
        raise ImportSheetError(file, [f'Could not find Stand: ({stand_name}) within file'])
    else:
        return rows


def read_excel(file, stand_name):
    rows = []
    wb = load_workbook(file, read_only=True, data_only=True)
    ws = wb.active

    table = ws.iter_rows(min_row=2, values_only=True)
    for row in table:
        if row[0].upper() == stand_name.upper():
            rows.append(row)

    if not rows:
        raise ImportSheetError(file, [f'Could not find Stand ({stand_name}) within file'])
    else:
        return rows


def error_check_quick(file, data_from_sheet):
    clean_rows = []
    error_message_bucket = []
    for i, row in enumerate(data_from_sheet):
        data = [row[0]]
        for j, col in enumerate(row[1:8], 1):
            error, val, message = SHEET_ROW_COL_CONV[j](col, i, j)
            if error:
                error_message_bucket.append(message)
            data.append(val)
        clean_rows.append(data)

    if error_message_bucket:
        raise ImportSheetError(file, error_message_bucket)
    else:
        return clean_rows


def error_check_full(file, data_from_sheet):
    clean_rows = []
    error_message_bucket = []
    for i, row in enumerate(data_from_sheet):
        data = [row[0]]
        # Main Data Check
        for j, col in enumerate(row[1:6], 1):
            error, val, message = SHEET_ROW_COL_CONV[j](col, i, j)
            if error:
                error_message_bucket.append(message)
            data.append(val)

        # Stump Height Check
        error, stem_height, message = is_err_num(row[6], i, 7, int, default=1)
        if error:
            error_message_bucket.append(message)

        # Logs Check
        logs = []
        for j in range(7, len(row), 4):
            log_length = row[j]
            if log_length != '' and log_length is not None:
                log = row[j: j + 4]
                clean_log = []
                for k, lg in enumerate(log):
                    error, val, message = SHEET_FULL_LOG_CONV[k](lg, i, j+k)
                    if error:
                        error_message_bucket.append(message)
                    clean_log.append(val)

                stem_height += clean_log[0] + 1
                # log_args = [Stem Height, Length, Grade, Defect]
                log_args = [stem_height] + clean_log[:3]
                stem_height += clean_log[-1]

                logs.append(log_args)

        data.append(logs)
        clean_rows.append(data)

    if error_message_bucket:
        raise ImportSheetError(file, error_message_bucket)
    else:
        return clean_rows


def get_plots(clean_data):
    plots = {}
    hdr = get_hdr(clean_data)
    for row in clean_data:
        plot = row[1]
        if plot not in plots:
            plots[plot] = []

        # total height check
        if not row[5]:
            row[5] = round((row[4] / 12) * hdr, 1)

        plots[plot].append(row[3:])

    return plots


def get_hdr(clean_data):
    return mean([row[5] / (row[4] / 12) for row in clean_data if row[5]])


def import_from_sheet(file, stand_name, cruise_type):
    ext = get_extension(file)
    if ext in ['.csv', '.xlsx']:
        if ext == '.csv':
            data = read_csv(file, stand_name)
        else:
            data = read_excel(file, stand_name)

        if cruise_type == 'q':
            clean = error_check_quick(file, data)
        else:
            clean = error_check_full(file, data)

        return get_plots(clean)
    else:
        raise ImportSheetError(file, [f'Not a valid sheet file with extension ({ext})', 'Need either .csv or .xlsx'])

