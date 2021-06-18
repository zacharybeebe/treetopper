from csv import reader
from openpyxl import load_workbook
from statistics import mean


def import_csv_quick(directory, stand_name):
    hdrs = []
    plots = {}
    with open(directory, 'r') as csv_file:
        csv_read = reader(csv_file)
        next(csv_read)
        for line in csv_read:
            if line[0] == stand_name:
                pnum = int(line[1])
                if pnum not in plots:
                    plots[pnum] = {}
                dbh = float(line[4])
                check_height = line[5]
                if check_height != '':
                    height = int(check_height)
                    hdrs.append(height / (dbh / 12))
                else:
                    height = check_height

                plots[pnum][int(line[2])] = {'species': line[3],
                                             'dbh': dbh,
                                             'height': height,
                                             'pref_log': int(line[6]),
                                             'min_log': int(line[7])}
            else:
                if plots:
                    break

    avg_hdr = mean(hdrs)
    return plots, avg_hdr


def import_csv_full(directory, stand_name):
    hdrs = []
    plots = {}
    with open(directory, 'r') as csv_file:
        csv_read = reader(csv_file)
        next(csv_read)
        for line in csv_read:
            if line[0] == stand_name:
                pnum = int(line[1])
                if pnum not in plots:
                    plots[pnum] = {}
                dbh = float(line[4])
                check_height = line[5]
                if check_height != '':
                    height = int(float(check_height))
                    hdrs.append(height / (dbh / 12))
                else:
                    height = check_height
                tree_num = int(line[2])
                plots[pnum][tree_num] = {'species': line[3],
                                         'dbh': dbh,
                                         'height': height,}
                logs = {}
                stem_height = int(line[6])
                for i in range(7, 40, 4):
                    if line[i] == '':
                        break
                    else:
                        length = int(line[i])
                        stem_height += length + 1
                        grade = line[i + 1].upper()
                        if line[i + 2] != '':
                            defect = int(line[i + 2])
                        else:
                            defect = 0
                        logs[(i // 4) - 1] = [stem_height, length, grade, defect]
                        if i != 40:
                            if line[i + 3] != '':
                                stem_height += int(line[i + 3])
                plots[pnum][tree_num]['logs'] = logs
            else:
                if plots:
                    break

    avg_hdr = mean(hdrs)
    return plots, avg_hdr


def import_excel_quick(directory, stand_name):
    hdrs = []
    plots = {}

    wb = load_workbook(directory, read_only=True, data_only=True)
    ws = wb.active

    table = ws.iter_rows(min_row=2, values_only=True)
    for line in table:
        if line[0] == stand_name:
            pnum = line[1]
            if pnum not in plots:
                plots[pnum] = {}
            dbh = line[4]
            height = line[5]
            if height:
                hdrs.append(height / (dbh / 12))

            plots[pnum][line[2]] = {'species': line[3],
                                    'dbh': dbh,
                                    'height': height,
                                    'pref_log': line[6],
                                    'min_log': line[7]}
        else:
            if plots:
                break

    avg_hdr = mean(hdrs)
    return plots, avg_hdr


def import_excel_full(directory, stand_name):
    hdrs = []
    plots = {}

    wb = load_workbook(directory, read_only=True, data_only=True)
    ws = wb.active

    table = ws.iter_rows(min_row=2, values_only=True)
    for line in table:
        if line[0] == stand_name:
            pnum = line[1]
            if pnum not in plots:
                plots[pnum] = {}
            dbh = line[4]
            height = line[5]
            if height:
                hdrs.append(height / (dbh / 12))
            tree_num = line[2]
            plots[pnum][tree_num] = {'species': line[3],
                                    'dbh': dbh,
                                    'height': height}
            logs = {}
            stem_height = line[6]
            for i in range(7, 40, 4):
                length = line[i]
                if not length:
                    break
                else:
                    stem_height += length + 1
                    grade = line[i+1].upper()
                    if line[i+2]:
                        defect = line[i+2]
                    else:
                        defect = 0
                    logs[(i // 4) - 1] = [stem_height, length, grade, defect]
                    if i != 40:
                        if line[i+3]:
                            stem_height += line[i+3]
            plots[pnum][tree_num]['logs'] = logs

        else:
            if plots:
                break

    avg_hdr = mean(hdrs)
    return plots, avg_hdr
